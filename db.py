import sqlite3
import sys
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class A4GDB:
    def __init__ (self, ATTFolderPath, KTFolderPath):

        self.PuertoRicoSA = ["PSE", "SJU" ]
        self.VirginIslandsSA = ["STT", "STX"]
        self.gateways = ['LAX', 'JFK', 'JFB', 'ATL', 'MIA', 'CVG']

        self.ATTpath = ATTFolderPath
        self.KTpath = KTFolderPath

        self.conn = sqlite3.connect('A4G.db')

        self.cur = self.conn.cursor()

        self.cur.execute("DROP TABLE IF EXISTS ZipCode")
        self.cur.execute("DROP TABLE IF EXISTS Route")
        self.cur.execute("DROP TABLE IF EXISTS Facility")
        self.cur.execute("DROP TABLE IF EXISTS Service_Area")
        self.conn.commit()

        try:
            self.cur.execute("""CREATE TABLE Service_Area (
                             SA CHAR(3) NOT NULL,
                             CTRY CHAR(2), 
                             PRIMARY KEY (SA));""")
            
            self.cur.execute("""CREATE TABLE Facility (
                             FAC CHAR(3) NOT NULL, 
                             SA CHAR(3), 
                             PRIMARY KEY(FAC),
                             FOREIGN KEY(SA) REFERENCES Service_Area(SA));""")

            self.cur.execute("""CREATE TABLE Route (
                             Rt VARCHAR(10) NOT NULL,
                             FAC CHAR(3),
                             FOREIGN KEY(FAC) REFERENCES Facility(FAC));""")
            
            self.cur.execute("""CREATE TABLE ZipCode (
                             Zip VARCHAR(5) NOT NULL,
                             Rt VARCHAR(10),
                             FOREIGN KEY(Rt) REFERENCES Route(Rt));""")
        except Exception as e:
            print(f"An error occured: {e}")


    def kt_files(self):
        for filename in os.listdir(self.KTpath):
            if filename.endswith('xlsx'):
                file_path = os.path.join(self.KTpath, filename)
                self.load_kt(file_path)
        return
    
    def load_kt(self, file):
        wb = load_workbook(file)
        ws = wb['TacticalTours']
        serviceArea_cache = ["None"]
        facility_cache = ["None"]

        for row in ws.iter_rows(2, ws.max_row + 1):
            route = str(row[0].value).strip()
            serviceArea = str(row[9].value).strip()
            facility = str(row[11].value).strip()
            
            if serviceArea in self.PuertoRicoSA:
                if serviceArea != "AAA":
                    if serviceArea not in serviceArea_cache:
                        print(serviceArea)
                        self.cur.execute(f"INSERT INTO Service_Area (SA, CTRY) VALUES ('{serviceArea}', 'PR');")
                        serviceArea_cache.append(serviceArea)

                    if facility not in facility_cache:
                        print(facility)
                        self.cur.execute(f"INSERT INTO Facility (FAC, SA) VALUES ('{facility}','{serviceArea}');")
                        facility_cache.append(facility)
                    if route != "None":
                        print(route)
                        self.cur.execute(f"INSERT INTO Route (Rt, FAC) VALUES('{route}', '{facility}');")
            elif serviceArea in self.VirginIslandsSA:
                if serviceArea != "AAA":
                    if serviceArea not in serviceArea_cache:
                        print(serviceArea)
                        self.cur.execute(f"INSERT INTO Service_Area (SA, CTRY) VALUES ('{serviceArea}', 'VI');")
                        serviceArea_cache.append(serviceArea)

                    if facility not in facility_cache:
                        print(facility)
                        self.cur.execute(f"INSERT INTO Facility (FAC, SA) VALUES ('{facility}','{serviceArea}');")
                        facility_cache.append(facility)
                    if route != "None":
                        print(route)
                        self.cur.execute(f"INSERT INTO Route (Rt, FAC) VALUES('{route}', '{facility}');")

            
            elif serviceArea not in self.gateways:
                if serviceArea != "AAA":
                    if serviceArea not in serviceArea_cache:
                        print(serviceArea)
                        self.cur.execute(f"INSERT INTO Service_Area (SA, CTRY) VALUES ('{serviceArea}', 'US');")
                        serviceArea_cache.append(serviceArea)

                    if facility not in facility_cache:
                        print(facility)
                        self.cur.execute(f"INSERT INTO Facility (FAC, SA) VALUES ('{facility}','{serviceArea}');")
                        facility_cache.append(facility)
                    if route != "None":
                        print(route)
                        self.cur.execute(f"INSERT INTO Route (Rt, FAC) VALUES('{route}', '{facility}');")
            
        self.conn.commit()
        return

    def att_files(self):
        for filename in os.listdir(self.ATTpath):
            if filename.endswith('xlsx'):
                file_path = os.path.join(self.ATTpath, filename)
                self.load_att(file_path)
        return

    def load_att(self, file):
        wb = load_workbook(file)
        print( wb.sheetnames)
        ws = wb['ATTPostalCode_SP1']
        route_cache = []

        for row in ws.iter_rows(2, ws.max_row + 1):
            if row[0].value is None or row[1].value is None or row[2].value is None:
                continue

            route = str(row[0].value).strip()
            code1 = int(row[1].value)
            code2 = int(row[2].value)

            try:
                for code in range(code1, code2 + 1):
                    pCode = str(code).zfill(5)
                    self.cur.execute(
                        "INSERT INTO ZipCode (Zip, Rt) VALUES (?, ?)", (pCode, route)
                    )
            except Exception as e:
                print(f"Route Doesn't Exist: {e}")

        self.conn.commit()



    def main(self):
        self.kt_files()
        self.att_files()



if __name__ == "__main__":
    att = "C:\\POcodeBot\\ATT Files"
    kt = "C:\\POcodeBot\\KT Files"
    db = A4GDB(att, kt)
    db.main()


 
